from __future__ import annotations

from typing import List, Optional
import openpyxl
import logging

# Import functions from the original Features.py file (renamed to features.py)
from features import ISI_time, StartEndFreq


def compute_basic_features(
    file_path: str,
    signal_vec: List,
    siz: int,
    mother: List,
    name: List,
    age: List,
    session: List,
    rec_num: List,
    mother_syl: List,
    name_syl: List,
    age_syl: List,
    session_syl: List,
    rec_num_syl: List,
    start_syl: List,
    end_syl: List,
    rate: int,
    logger: Optional[logging.Logger] = None,
) -> str:
    """
    Compute basic features (ISI time and start/end frequencies) and write to Excel.
    
    This function:
    1. Computes ISI (Inter-Syllable Interval) time
    2. Computes start and end frequencies for each syllable
    3. Writes the results to the Excel file as new columns
    
    Args:
        file_path: Path to the segmentation Excel file (will be updated)
        signal_vec: List of audio signal arrays
        siz: Number of recordings
        mother, name, age, session, rec_num: Metadata lists for recordings
        mother_syl, name_syl, age_syl, session_syl, rec_num_syl: Metadata lists for syllables
        start_syl, end_syl: Start and end times for each syllable
        rate: Sampling rate
        logger: Optional logger instance for logging
    
    Returns:
        Path to the updated Excel file
    """
    if logger:
        logger.info("Computing basic features: ISI time and start/end frequencies")
    
    # Compute features using original functions
    ISI = ISI_time(rec_num_syl, start_syl, end_syl)
    startF, endF = StartEndFreq(
        signal_vec, siz, mother, name, age, session, rec_num,
        mother_syl, name_syl, age_syl, session_syl, rec_num_syl,
        start_syl, end_syl, rate
    )
    
    # Write to Excel file
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.worksheets[0]
    
    # Find the last column and insert 3 new columns after it
    last_column = worksheet.max_column
    first_new_column = last_column + 1
    worksheet.insert_cols(first_new_column, 3)
    
    # Column names for the new columns
    column_names = ['ISI_time', 'Start Point (Hz)', 'End Point (Hz)']
    
    # Write column headers
    for col_idx, col_name in enumerate(column_names, start=first_new_column):
        worksheet.cell(row=1, column=col_idx).value = col_name
    
    # Write values (starting from row 2, row 1 is headers)
    for row_idx in range(len(ISI)):
        worksheet.cell(row=row_idx + 2, column=first_new_column).value = ISI[row_idx]
        worksheet.cell(row=row_idx + 2, column=first_new_column + 1).value = startF[row_idx]
        worksheet.cell(row=row_idx + 2, column=first_new_column + 2).value = endF[row_idx]
    
    workbook.save(file_path)
    
    if logger:
        logger.info(f"Basic features computed and written to {file_path}")
    
    return file_path

