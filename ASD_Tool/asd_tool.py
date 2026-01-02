import numpy as np
import pandas as pd
import os
import librosa
import logging
import glob
import openpyxl
from openpyxl import Workbook
from pytictoc import TicToc
from audio_feature_extraction_reduction_by_recording import *
from pipeline.utils import (
    setup_logger,
    list_metadata_files,
    is_already_processed,
    is_segmentation_file_exist,
    parse_args,
    get_files_to_process,
)
from pipeline.steps import prepare_recording_metadata, run_segmentation, read_segmentation_results, compute_basic_features


##################################################
#### 1: setup & file selection
##################################################
logger = setup_logger()

input_files = list_metadata_files("metadata")
logger.info(f"Found {len(input_files)} metadata file(s): {input_files}")

# Parse CLI arguments to check if user specified a single file to process
args = parse_args()
# Determine which files to process: all files or single file if --metadata-file was provided
# This function validates the file exists and logs the selection
files_to_process = get_files_to_process(input_files, args.metadata_file, logger)

for file_name in files_to_process:
  try:
    t = TicToc() #create instance of class
    t.tic() #Start timer
    logger.info(f"Starting file: {file_name}")

    # Skip files with existing outputs (xlsx/csv/npy) to resume safely and avoid unnecessary reprocessing
    if is_already_processed(file_name, "outputs"):
      logger.info(f"Skipping file (already processed): {file_name}")
      continue

    ##################################################
    #### 2: segmentation
    ##################################################
    # Load metadata + audio recordings
    (
        year,
        mother, matgen, name, sex, pupgen, age, session, rec_num,
        SignalVec, signal_name, rate, missing_count,
    ) = prepare_recording_metadata(
        file_name=file_name,
        metadata_dir="metadata",
        recordings_root="USV_Recordings",
        sr=250000,
        logger=logger,
    )
    siz = len(SignalVec)

    # Check if segmentation already exists - if so, skip segmentation step
    if is_segmentation_file_exist(file_name, "outputs"):
      logger.info(f"Segmentation already exists for {file_name}, skipping segmentation step")
      output_xlsx = f'outputs/{file_name}'
    else:
      # Run segmentation: process recordings, detect syllables, save to Excel
      output_xlsx = run_segmentation(
          file_name=file_name,
          SignalVec=SignalVec,
          signal_name=signal_name,
          rate=rate,
          mother=mother,
          matgen=matgen,
          name=name,
          sex=sex,
          pupgen=pupgen,
          age=age,
          session=session,
          rec_num=rec_num,
          missing_count=missing_count,
          logger=logger,
      )

    ##################################################
    #### 3: basic features (ISI time + start/end frequencies)
    ##################################################
    # Read segmentation results from Excel file (using column names)
    (
        motherSyl, matgenSyl, nameSyl, sexSyl, pupgenSyl,
        ageSyl, sessionSyl, rec_numSyl, startSyl, endSyl,
    ) = read_segmentation_results(f'outputs/{file_name}', logger=logger)

    # Compute basic features and add 3 columns to Excel: 'ISI_time', 'Start Point (Hz)', 'End Point (Hz)'
    compute_basic_features(
        file_path=f'outputs/{file_name}',
        signal_vec=SignalVec,
        siz=siz,
        mother=mother,
        name=name,
        age=age,
        session=session,
        rec_num=rec_num,
        mother_syl=motherSyl,
        name_syl=nameSyl,
        age_syl=ageSyl,
        session_syl=sessionSyl,
        rec_num_syl=rec_numSyl,
        start_syl=startSyl,
        end_syl=endSyl,
        rate=rate,
        logger=logger,
    )

    ##################################################
    #### 4: classification
    ##################################################
    logger.info("Classification started")
    from statistics_generator import *

    model_path = 'ASD_tool/model_weights.h6'
    # model = keras.models.load_model(model_path, custom_objects={'KerasLayer':hub.KerasLayer})
    model = keras.models.load_model(model_path)

    samples = Syl_Class_Vec(year, model,ageSyl,matgenSyl,pupgenSyl,motherSyl,nameSyl,sexSyl,sessionSyl,rec_numSyl,startSyl,endSyl)
    logger.debug(f"Samples: {samples}")
    output_npy = f"outputs/{file_name.split('.')[0]}.npy"
    np.save(output_npy, samples)

    syl_num = []
    for i in range(len(samples)):
      for j in range(len(samples[i].syls)):
        if np.max(samples[i].syls[j])<0.5:
          temp = 10
        else:
          temp = np.argmax(samples[i].syls[j])
        samples[i].syls[j] = []
        samples[i].syls[j] = temp
        syl_num.append(samples[i].syls[j])
        logger.debug(f"Syllable number: {samples[i].syls[j]}")


    y = 2
    workbook = openpyxl.load_workbook(f'outputs/{file_name}')
    worksheet = workbook.worksheets[0]
    worksheet.insert_cols(16)
    cell_title = worksheet.cell(row=1, column=16)
    cell_title.value = 'Syllable number'
    for x in range(len(syl_num)):
        cell_to_write = worksheet.cell(row=y, column=16)
        cell_to_write.value = syl_num[x]
        y += 1
    workbook.save(f'outputs/{file_name}')
    logger.info(f"Classification finished (syllables={len(syl_num)})")

    logger.info("Feature extraction started")

    dataset = pd.read_excel(f'outputs/{file_name}')

    # Extract only the relevant columns / features
    X = dataset[["Name", "Day", "Session", "Start Point (Hz)", "End Point (Hz)", "Duration (time)", "Syllable number", "Recording Number", "Mother Genotype", "Sex", "ISI_time", "Offspring Genotype"]]

    mouse_final_data = feature_extraction(X)
    # Export data to CSV file for further use
    output_csv = f"outputs/{file_name.split('.')[0]}.csv"
    np.savetxt(output_csv, X=mouse_final_data, delimiter=",")

    logger.info(f"Exported: {output_xlsx}, {output_csv}, {output_npy}")
    logger.info(f"Finished processing file: {file_name}")
    t.toc() #Time elapsed since t.tic()
  except Exception as e:
    logger.exception(f"Error processing file {file_name}: {e}")
    raise



##################################################
#### 5: aggregation (all files)
##################################################
def extract_features(dir):
  try:
    logger.info("Aggregating features from all processed files")
    # Extract all files with xlsx extension
    all_files = glob.glob(os.path.join('outputs' , "*.xlsx"))
    logger.info(f"Found {len(all_files)} processed file(s)")
    # Read and combine all input files
    dataset = pd.concat((pd.read_excel(f) for f in all_files), ignore_index=True)
    # Add Strain column based on year in path
    dataset["Strain"] = [1 if int(x.split('/')[1]) == 2022 else 2 for x in dataset['Path']]
    dataset.to_excel(f"{dir}/all_data.xlsx")
    # Extract only the relevant columns / features
    dataset = dataset[["Name", "Day", "Session", "Start Point (Hz)", "End Point (Hz)", "Duration (time)", "Syllable number", "Recording Number", "Mother Genotype", "Sex", "ISI_time", "Offspring Genotype", "Strain"]]
    # Extract features
    mouse_final_data = feature_extraction(dataset)
    # Save the output file
    output_path = f"{dir}/all_data.csv"
    np.savetxt(output_path, X=mouse_final_data, delimiter=",")
    logger.info(f"Finished aggregating features: {dir}/all_data.xlsx, {output_path}")
  except Exception as e:
    logger.exception(f"Error in extract_features: {e}")
    raise

if __name__ == "__main__":
  extract_features(dir='outputs')
